#!/usr/bin/env python
"""
Label Import Script - Load Labels from Illumine-i X LabelFriday Excel File
Imports label data with UPC codes, dimensions, colors, descriptions, and images.

The Excel file structure:
- Column A: UPC/ID (e.g., LF-001)
- Column B: Label name (e.g., WRITE IN LABEL)
- Column C: Description (e.g., Write in DC label)
- Column D: Length (e.g., 4")
- Column E: Width (e.g., 1")
- Column F: Image URL
- Column G-H: Automation Used, PID (skipped)
- Column I: Background color (e.g., Orange, Blue)
- Column J: Text color (e.g., Black, White)

Usage:
    python scripts/import/import_labels.py

Note: Ensure you are in the project root directory when running this script.
"""

import os
import sys
import re
from pathlib import Path
from typing import Optional, Dict

# Add project root to path for absolute imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session

from app.core.database import engine, SessionLocal
from app.models.label import Label


# Color name to hex code mapping
COLOR_MAP = {
    "orange": "#FFA500",
    "black": "#000000",
    "white": "#FFFFFF",
    "red": "#FF0000",
    "blue": "#0000FF",
    "green": "#008000",
    "yellow": "#FFFF00",
    "purple": "#800080",
    "pink": "#FFC0CB",
    "brown": "#A52A2A",
    "gray": "#808080",
    "grey": "#808080",
    "cyan": "#00FFFF",
    "magenta": "#FF00FF",
    "silver": "#C0C0C0",
    "gold": "#FFD700",
    "navy": "#000080",
    "teal": "#008080",
    "olive": "#808000",
    "maroon": "#800000",
}


class LabelImporter:
    """
    Handles importing label data from Excel to PostgreSQL.
    Converts color names to hex codes and parses dimensions.
    """

    def __init__(self):
        self.db: Session = SessionLocal()
        self.stats = {
            "labels_created": 0,
            "labels_skipped": 0,
            "labels_errors": 0,
            "errors": []
        }

    def get_excel_path(self, filename: str) -> str:
        """
        Build absolute path to Excel file in data/master/ directory.
        """
        possible_paths = [
            PROJECT_ROOT / "data" / "master" / filename,
            PROJECT_ROOT / "app" / "data" / "master" / filename,
            PROJECT_ROOT / "data" / filename,
            PROJECT_ROOT / "app" / "data" / filename,
            Path.cwd() / "data" / "master" / filename,
            Path.cwd() / "app" / "data" / "master" / filename,
            Path.cwd() / "data" / filename,
            Path.cwd() / "app" / "data" / filename,
        ]

        for path in possible_paths:
            if path.exists():
                print(f"✓ Found Excel file: {path}")
                return str(path)

        raise FileNotFoundError(
            f"Excel file '{filename}' not found at:\n" +
            "\n".join(f"  - {p}" for p in possible_paths)
        )

    def load_workbook_safely(self, filepath: str):
        """Load Excel workbook with error handling."""
        try:
            return openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            raise Exception(f"Failed to load {filepath}: {str(e)}")

    def parse_dimension(self, dimension_str: str) -> Optional[int]:
        """
        Parse dimension string like '4"' or ' 1"' to integer.
        Returns the numeric value, or None if parsing fails.
        """
        if not dimension_str or not str(dimension_str).strip():
            return None
        
        try:
            # Remove quotes, spaces, and convert to float
            cleaned = str(dimension_str).strip().replace('"', '').strip()
            # Try to extract just the number
            match = re.search(r'[\d.]+', cleaned)
            if match:
                value = float(match.group())
                return int(value)  # Convert to int for storage
        except (ValueError, AttributeError):
            pass
        
        return None

    def color_name_to_hex(self, color_name: str) -> Optional[str]:
        """
        Convert color name (e.g., "Orange") to hex code (e.g., "#FFA500").
        Returns hex code if found, otherwise returns None.
        """
        if not color_name or not str(color_name).strip():
            return None
        
        # Check for error values
        if str(color_name).upper() in ["#VALUE!", "#N/A", "N/A", ""]:
            return None
        
        try:
            color_key = str(color_name).strip().lower()
            
            # Check if it's already a hex code
            if color_key.startswith('#') and len(color_key) in [7, 9]:
                return color_key
            
            # Look up in map
            if color_key in COLOR_MAP:
                return COLOR_MAP[color_key]
        except:
            pass
        
        return None

    def label_exists(self, upc_code: str) -> bool:
        """Check if a label with this UPC code already exists."""
        if not upc_code or not str(upc_code).strip():
            return False
        
        existing = self.db.query(Label).filter_by(upc_code=str(upc_code).strip()).first()
        return existing is not None

    def import_labels(self):
        """
        Import labels from the "Illumine-i X LabelFriday master data.xlsx" file.
        """
        print("\n" + "="*70)
        print("IMPORTING LABELS FROM ILLUMINE-I X LABELFRIDAY")
        print("="*70)
        
        try:
            filepath = self.get_excel_path("Illumine-i X LabelFriday master data.xlsx")
            print(f"\nLoading workbook: {filepath}")
            
            wb = self.load_workbook_safely(filepath)
            ws = wb["Overall"]
            
            print(f"✓ Found sheet: 'Overall'")
            print(f"  Dimensions: {ws.dimensions}")
            
            print("\n--- Processing Labels ---")
            
            label_count = 0
            batch_count = 0
            label_number_counter = 1000  # Start numbering from 1000
            
            # Process data rows (starting from row 2, skipping header)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 10:
                    continue
                
                try:
                    # Extract columns
                    upc_code = row[0]  # Column A
                    label_name = row[1]  # Column B
                    description = row[2]  # Column C
                    length_str = row[3]  # Column D
                    width_str = row[4]  # Column E
                    image_url = row[5]  # Column F
                    background_color_name = row[8]  # Column I
                    text_color_name = row[9]  # Column J
                    
                    # Skip if no UPC or name
                    if not upc_code or not label_name:
                        continue
                    
                    upc_code = str(upc_code).strip()
                    label_name = str(label_name).strip()
                    
                    # Skip #VALUE! and #N/A entries
                    if upc_code.startswith("#") or label_name.startswith("#"):
                        continue
                    
                    # Check for duplicates
                    if self.label_exists(upc_code):
                        self.stats["labels_skipped"] += 1
                        continue
                    
                    # Parse dimensions
                    length = self.parse_dimension(length_str)
                    width = self.parse_dimension(width_str)
                    
                    # Convert colors
                    bg_color = self.color_name_to_hex(background_color_name)
                    text_color = self.color_name_to_hex(text_color_name)
                    
                    # Handle image URL (skip #VALUE! entries)
                    image_url_str = None
                    if image_url and not str(image_url).startswith("#"):
                        image_url_str = str(image_url).strip()
                    
                    # Parse description
                    description_str = None
                    if description and str(description).strip() and not str(description).startswith("#"):
                        description_str = str(description).strip()
                    
                    # Create Label record
                    label = Label(
                        upc_code=upc_code,
                        name=label_name,
                        label_number=f"LN-{label_number_counter}",
                        label_name=description_str,
                        description=description_str,
                        length=length,
                        width=width,
                        image_url=image_url_str,
                        background_color=bg_color,
                        text_color=text_color,
                        is_active=True
                    )
                    label_number_counter += 1
                    
                    self.db.add(label)
                    self.stats["labels_created"] += 1
                    label_count += 1
                    
                    if label_count % 50 == 0:
                        print(f"  ✓ Processed {label_count} labels...")
                    
                    # Commit every 100 labels
                    if label_count % 100 == 0:
                        self.db.commit()
                        batch_count += 1
                        print(f"  ✓ Committed batch {batch_count} ({label_count} labels)")
                
                except Exception as e:
                    self.db.rollback()
                    self.stats["labels_errors"] += 1
                    error_msg = f"Row {row_idx}: {str(e)}"
                    self.stats["errors"].append(error_msg)
            
            # Final commit
            if label_count % 100 != 0:
                self.db.commit()
                if label_count > 0:
                    print(f"  ✓ Final commit ({label_count} labels)")
            
            print("\n✓ Excel file processed successfully")
        
        except Exception as e:
            self.db.rollback()
            print(f"\n✗ ERROR: {str(e)}")
            raise

    def print_summary(self):
        """Print import summary statistics."""
        print("\n" + "="*70)
        print("LABEL IMPORT SUMMARY")
        print("="*70)
        
        print(f"\nLabels Created: {self.stats['labels_created']}")
        print(f"Labels Skipped (Duplicates): {self.stats['labels_skipped']}")
        print(f"Labels Errors: {self.stats['labels_errors']}")
        
        if self.stats["errors"]:
            print("\nFirst 5 Errors:")
            for error in self.stats["errors"][:5]:
                print(f"  - {error}")
            if len(self.stats["errors"]) > 5:
                print(f"  ... and {len(self.stats['errors']) - 5} more errors")
        
        print("\n" + "="*70)
        print("✓ LABEL IMPORT COMPLETE!")
        print("="*70)

    def run(self):
        """Run the complete import process."""
        try:
            self.import_labels()
            self.print_summary()
        except Exception as e:
            print(f"\n✗ IMPORT FAILED: {str(e)}")
            self.db.rollback()
            raise
        finally:
            self.db.close()


def main():
    """Main entry point for the script."""
    importer = LabelImporter()
    importer.run()


if __name__ == "__main__":
    main()
