#!/usr/bin/env python
"""
Regulatory Data Import Script - Load Country, State, AHJ, and Utility Mappings
This script imports data from Regulatory data.xlsx containing:
1. Country List - Countries with ISO codes
2. State List - USA states with abbreviations
3. AHJ - State List - AHJs mapped to states
4. AHJ - Utility mapping - Utilities mapped to AHJs

Usage:
    python scripts/import/import_regulatory_data.py

Note: Ensure you are in the project root directory when running this script.
"""

import os
import sys
from pathlib import Path

# Add project root to path for absolute imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session
from app.core.database import engine, SessionLocal
from app.models import Country, State, AHJ, Utility


class RegulatoryDataImporter:
    """
    Handles importing regulatory data from Excel files to PostgreSQL.
    - Validates data before inserting
    - Handles duplicates gracefully
    - Provides detailed logging and statistics
    """

    def __init__(self):
        self.db = SessionLocal()
        self.stats = {
            "countries_created": 0,
            "countries_skipped": 0,
            "states_created": 0,
            "states_skipped": 0,
            "ahjs_created": 0,
            "ahjs_skipped": 0,
            "utilities_created": 0,
            "utilities_skipped": 0,
            "errors": []
        }

    def get_excel_path(self) -> str:
        """Build absolute path to Regulatory data.xlsx"""
        possible_paths = [
            PROJECT_ROOT / "data" / "master" / "Regulatory data.xlsx",
            PROJECT_ROOT / "app" / "data" / "master" / "Regulatory data.xlsx",
            Path.cwd() / "data" / "master" / "Regulatory data.xlsx",
            Path.cwd() / "app" / "data" / "master" / "Regulatory data.xlsx",
        ]

        for path in possible_paths:
            if path.exists():
                return str(path)

        raise FileNotFoundError(
            f"Excel file 'Regulatory data.xlsx' not found at:\n" +
            "\n".join(f"  - {p}" for p in possible_paths)
        )

    def load_workbook_safely(self, filepath: str):
        """Load Excel workbook safely"""
        try:
            return openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"ERROR: Failed to load Excel file: {e}")
            sys.exit(1)

    def import_countries(self, wb):
        """Import countries from 'Country List' sheet"""
        print("\n" + "="*60)
        print("IMPORTING COUNTRIES")
        print("="*60)
        
        try:
            ws = wb["Country List"]
        except KeyError:
            print("ERROR: 'Country List' sheet not found in workbook")
            return

        rows = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue

            rows += 1
            country_name = row[0]
            iso2 = row[1]

            if not country_name or not iso2:
                continue

            try:
                iso2_clean = str(iso2).strip().upper()[:2]
                iso3_from_sheet = row[2] if len(row) > 2 else None
                iso3_clean = (
                    str(iso3_from_sheet).strip().upper()[:3]
                    if iso3_from_sheet and str(iso3_from_sheet).strip()
                    else f"{iso2_clean}X"
                )

                # Check if country already exists
                existing = self.db.query(Country).filter_by(iso2=iso2_clean).first()
                if existing:
                    self.stats["countries_skipped"] += 1
                    continue

                # Create new country
                country = Country(
                    name=country_name,
                    iso2=iso2_clean,
                    iso3=iso3_clean,
                    calling_code=None,
                    currency_code=None
                )
                self.db.add(country)
                self.db.commit()
                self.stats["countries_created"] += 1
                print(f"[OK] Created: {country_name} ({iso2_clean})")

            except Exception as e:
                self.db.rollback()
                error_msg = f"Row {idx}: {str(e)}"
                self.stats["errors"].append(error_msg)
                print(f"[ERROR] Error on row {idx}: {e}")

        print(f"\nCountries Summary:")
        print(f"  Created: {self.stats['countries_created']}")
        print(f"  Skipped: {self.stats['countries_skipped']}")
        print(f"  Total rows processed: {rows}")

    def import_states(self, wb):
        """Import USA states from 'State List - USA' sheet"""
        print("\n" + "="*60)
        print("IMPORTING STATES")
        print("="*60)
        
        try:
            ws = wb["State List - USA"]
        except KeyError:
            print("ERROR: 'State List - USA' sheet not found in workbook")
            return

        rows = 0
        # Get USA country
        usa_country = self.db.query(Country).filter(
            Country.iso2.ilike("us")
        ).first()
        if not usa_country:
            print("ERROR: USA country not found. Import countries first.")
            return

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue

            rows += 1
            state_name = row[0]
            abbreviation = row[1]

            if not state_name or not abbreviation:
                continue

            try:
                # Check if state already exists
                existing = self.db.query(State).filter_by(abbrev=abbreviation).first()
                if existing:
                    self.stats["states_skipped"] += 1
                    continue

                # Create new state
                state = State(
                    name=state_name,
                    abbrev=abbreviation,
                    country_id=usa_country.id,
                    fips_code=None,
                    region=None
                )
                self.db.add(state)
                self.db.commit()
                self.stats["states_created"] += 1
                print(f"[OK] Created: {state_name} ({abbreviation})")

            except Exception as e:
                self.db.rollback()
                error_msg = f"Row {idx}: {str(e)}"
                self.stats["errors"].append(error_msg)
                print(f"[ERROR] Error on row {idx}: {e}")

        print(f"\nStates Summary:")
        print(f"  Created: {self.stats['states_created']}")
        print(f"  Skipped: {self.stats['states_skipped']}")
        print(f"  Total rows processed: {rows}")

    def import_ahjs(self, wb):
        """Import AHJs from 'AHJ - State List' sheet"""
        print("\n" + "="*60)
        print("IMPORTING AHJs (AUTHORITIES HAVING JURISDICTION)")
        print("="*60)
        
        try:
            ws = wb["AHJ - State List"]
        except KeyError:
            print("ERROR: 'AHJ - State List' sheet not found in workbook")
            return

        rows = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue

            rows += 1
            ahj_name = row[0]
            state_name = row[1]

            if not ahj_name or not state_name:
                continue

            try:
                # Extract state abbreviation from AHJ name (e.g., "CA - City Name" -> "CA")
                state_abbrev = None
                if " - " in ahj_name:
                    state_abbrev = ahj_name.split(" - ")[0].strip()

                # Find the state by abbreviation first
                state = None
                if state_abbrev:
                    state = self.db.query(State).filter_by(abbrev=state_abbrev).first()
                
                # Fallback: look up by state name
                if not state and state_name:
                    state = self.db.query(State).filter_by(name=state_name).first()

                if not state:
                    print(f"[WARN] Warning Row {idx}: State '{state_abbrev or state_name}' not found, skipping AHJ '{ahj_name}'")
                    self.stats["ahjs_skipped"] += 1
                    continue

                # Check if AHJ already exists
                existing = self.db.query(AHJ).filter_by(
                    name=ahj_name,
                    state_id=state.id
                ).first()
                if existing:
                    self.stats["ahjs_skipped"] += 1
                    continue

                # Create new AHJ
                ahj = AHJ(
                    name=ahj_name,
                    ahj_name=ahj_name,
                    state_id=state.id,
                    county=None,
                    city=None,
                    guidelines=None,
                    fireset_back=None,
                    jurisdiction_type=None,
                    phone=None,
                    email=None,
                    website=None
                )
                self.db.add(ahj)
                self.db.commit()
                self.stats["ahjs_created"] += 1
                print(f"[OK] Created: {ahj_name} ({state_abbrev})")

            except Exception as e:
                self.db.rollback()
                error_msg = f"Row {idx}: {str(e)}"
                self.stats["errors"].append(error_msg)
                print(f"[ERROR] Error on row {idx}: {e}")

        print(f"\nAHJs Summary:")
        print(f"  Created: {self.stats['ahjs_created']}")
        print(f"  Skipped: {self.stats['ahjs_skipped']}")
        print(f"  Total rows processed: {rows}")

    def import_utility_ahj_mapping(self, wb):
        """Import Utility-AHJ mappings from 'AHJ - Utility mapping' sheet"""
        print("\n" + "="*60)
        print("IMPORTING UTILITY-AHJ MAPPINGS")
        print("="*60)
        
        try:
            ws = wb["AHJ - Utility mapping"]
        except KeyError:
            print("ERROR: 'AHJ - Utility mapping' sheet not found in workbook")
            return

        rows = 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip empty rows
                continue

            rows += 1
            utility_name = row[0]
            ahj_name = row[1]

            if not utility_name or not ahj_name:
                continue

            try:
                # Find the AHJ
                ahj = self.db.query(AHJ).filter_by(name=ahj_name).first()
                if not ahj:
                    print(f"[WARN] Warning Row {idx}: AHJ '{ahj_name}' not found, skipping utility '{utility_name}'")
                    self.stats["utilities_skipped"] += 1
                    continue

                # Check if utility already exists
                existing = self.db.query(Utility).filter_by(
                    name=utility_name,
                    state_id=ahj.state_id
                ).first()
                if existing:
                    self.stats["utilities_skipped"] += 1
                    continue

                # Create new utility
                utility = Utility(
                    name=utility_name,
                    utility_name=utility_name,
                    state_id=ahj.state_id,
                    ahj_id=ahj.id,
                    eia_id=None,
                    utility_type=None,
                    service_territory=None,
                    phone=None,
                    website=None,
                    requirements=None,
                    response_type=None
                )
                self.db.add(utility)
                self.db.commit()
                self.stats["utilities_created"] += 1
                print(f"[OK] Created: {utility_name} (linked to {ahj_name})")

            except Exception as e:
                self.db.rollback()
                error_msg = f"Row {idx}: {str(e)}"
                self.stats["errors"].append(error_msg)
                print(f"[ERROR] Error on row {idx}: {e}")

        print(f"\nUtilities Summary:")
        print(f"  Created: {self.stats['utilities_created']}")
        print(f"  Skipped: {self.stats['utilities_skipped']}")
        print(f"  Total rows processed: {rows}")

    def run(self):
        """Run the complete import process"""
        print("\n" + "="*60)
        print("REGULATORY DATA IMPORT")
        print("="*60 + "\n")

        try:
            filepath = self.get_excel_path()
            print(f"Loading Excel file: {filepath}\n")
            wb = self.load_workbook_safely(filepath)

            # Import in order (countries first, then states, then AHJs)
            self.import_countries(wb)
            self.import_states(wb)
            self.import_ahjs(wb)
            self.import_utility_ahj_mapping(wb)

            # Print final summary
            self._print_summary()

        except Exception as e:
            print(f"\nFATAL ERROR: {e}")
            sys.exit(1)
        finally:
            self.db.close()

    def _print_summary(self):
        """Print import summary"""
        print("\n" + "="*60)
        print("IMPORT SUMMARY")
        print("="*60)
        print(f"Countries Created: {self.stats['countries_created']}")
        print(f"States Created: {self.stats['states_created']}")
        print(f"AHJs Created: {self.stats['ahjs_created']}")
        print(f"Utilities Created: {self.stats['utilities_created']}")
        print(f"\nTotal Records Created: {sum([self.stats['countries_created'], self.stats['states_created'], self.stats['ahjs_created'], self.stats['utilities_created']])}")
        print(f"Total Records Skipped: {sum([self.stats['countries_skipped'], self.stats['states_skipped'], self.stats['ahjs_skipped'], self.stats['utilities_skipped']])}")

        if self.stats["errors"]:
            print(f"\nErrors Encountered: {len(self.stats['errors'])}")
            for error in self.stats["errors"][:10]:  # Show first 10 errors
                print(f"  - {error}")
            if len(self.stats["errors"]) > 10:
                print(f"  ... and {len(self.stats['errors']) - 10} more errors")

        print("\n" + "="*60)
        print("IMPORT COMPLETE!")
        print("="*60 + "\n")


if __name__ == "__main__":
    importer = RegulatoryDataImporter()
    importer.run()
