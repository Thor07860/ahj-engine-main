#!/usr/bin/env python
"""
Equipment, Category & CombinationMapper Import Script
Imports equipment types and categories from Excel, then creates CombinationMapper entries.

The Excel file structure:
- Column A: Equipment Category (e.g., "Module", "Inverter", "Racking")
- Column B: Equipment Type (e.g., "DC Module", "AC Module", "Microinverter")

Script creates:
1. Category records from column A
2. Equipment records from column B
3. CombinationMapper entries linking equipment + category

Usage:
    python scripts/import/import_equipment_categories.py

Note: Ensure you are in the project root directory when running this script.
"""

import os
import sys
from pathlib import Path
from typing import Dict, List, Tuple

# Add project root to path for absolute imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session

from app.core.database import engine, SessionLocal
from app.models.category import Category
from app.models.equipment import Equipment
from app.models.combination_mapper import CombinationMapper


class EquipmentCategoryImporter:
    """
    Handles importing equipment and category data from Excel,
    then creates CombinationMapper entries linking them.
    """

    def __init__(self):
        self.db: Session = SessionLocal()
        self.stats = {
            "categories_created": 0,
            "categories_skipped": 0,
            "equipment_created": 0,
            "equipment_skipped": 0,
            "mappings_created": 0,
            "mappings_skipped": 0,
            "errors": []
        }
        
        # Cache for lookups
        self.categories_cache: Dict[str, int] = {}
        self.equipment_cache: Dict[str, int] = {}

    def get_excel_path(self, filename: str) -> str:
        """
        Build absolute path to Excel file in data/master/ directory.
        """
        possible_paths = [
            PROJECT_ROOT / "data" / "master" / filename,
            PROJECT_ROOT / "app" / "data" / "master" / filename,
            PROJECT_ROOT / "data" / filename,
            PROJECT_ROOT / "app" / "data" / filename,
            Path.cwd() / "data" / "master" / filename,
            Path.cwd() / "app" / "data" / "master" / filename,
            Path.cwd() / "data" / filename,
            Path.cwd() / "app" / "data" / filename,
        ]

        for path in possible_paths:
            if path.exists():
                print(f"✓ Found Excel file: {path}")
                return str(path)

        raise FileNotFoundError(
            f"Excel file '{filename}' not found at:\n" +
            "\n".join(f"  - {p}" for p in possible_paths)
        )

    def load_workbook_safely(self, filepath: str):
        """Load Excel workbook with error handling."""
        try:
            return openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            raise Exception(f"Failed to load {filepath}: {str(e)}")

    def get_or_create_category(self, category_name: str) -> int:
        """
        Get or create a Category by name.
        Returns the ID.
        """
        if not category_name or not str(category_name).strip():
            return None
        
        category_name = str(category_name).strip()
        
        # Check cache first
        if category_name in self.categories_cache:
            return self.categories_cache[category_name]
        
        # Query database
        existing = self.db.query(Category).filter_by(name=category_name).first()
        
        if not existing:
            category = Category(name=category_name)
            self.db.add(category)
            self.db.flush()
            category_id = category.id
            self.stats["categories_created"] += 1
            print(f"  ✓ Created Category: {category_name}")
        else:
            category_id = existing.id
            self.stats["categories_skipped"] += 1
        
        # Cache it
        self.categories_cache[category_name] = category_id
        return category_id

    def get_or_create_equipment(self, equipment_name: str) -> int:
        """
        Get or create Equipment by name.
        Returns the ID.
        """
        if not equipment_name or not str(equipment_name).strip():
            return None
        
        equipment_name = str(equipment_name).strip()
        
        # Check cache first
        if equipment_name in self.equipment_cache:
            return self.equipment_cache[equipment_name]
        
        # Query database
        existing = self.db.query(Equipment).filter_by(name=equipment_name).first()
        
        if not existing:
            equipment = Equipment(name=equipment_name)
            self.db.add(equipment)
            self.db.flush()
            equipment_id = equipment.id
            self.stats["equipment_created"] += 1
            print(f"  ✓ Created Equipment: {equipment_name}")
        else:
            equipment_id = existing.id
            self.stats["equipment_skipped"] += 1
        
        # Cache it
        self.equipment_cache[equipment_name] = equipment_id
        return equipment_id

    def combination_mapper_exists(self, category_id: int, equipment_id: int) -> bool:
        """
        Check if a CombinationMapper already exists for this category + equipment combination.
        """
        if not category_id or not equipment_id:
            return False
        
        existing = self.db.query(CombinationMapper).filter(
            CombinationMapper.category_id == category_id,
            CombinationMapper.equipment_id == equipment_id
        ).first()
        
        return existing is not None

    def import_equipment_and_categories(self):
        """
        Import equipment and categories from Excel, then create CombinationMapper entries.
        """
        print("\n" + "="*70)
        print("IMPORTING EQUIPMENT, CATEGORIES & COMBINATION MAPPERS")
        print("="*70)
        
        try:
            filepath = self.get_excel_path("Equipment Category and types.xlsx")
            print(f"\nLoading workbook: {filepath}")
            
            wb = self.load_workbook_safely(filepath)
            ws = wb["Sheet1"]
            
            print(f"✓ Found sheet: 'Sheet1'")
            print(f"  Dimensions: {ws.dimensions}")
            
            # Read all data rows (starting from row 2, skipping header)
            print("\n--- Processing Equipment & Categories ---")
            
            mapper_count = 0
            batch_count = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if len(row) < 2:
                    continue
                
                category_name = row[0]
                equipment_name = row[1]
                
                # Skip empty rows
                if not category_name or not equipment_name:
                    continue
                
                try:
                    category_name = str(category_name).strip()
                    equipment_name = str(equipment_name).strip()
                    
                    # Get or create category
                    category_id = self.get_or_create_category(category_name)
                    
                    # Get or create equipment
                    equipment_id = self.get_or_create_equipment(equipment_name)
                    
                    # Check if mapping already exists
                    if self.combination_mapper_exists(category_id, equipment_id):
                        self.stats["mappings_skipped"] += 1
                        continue
                    
                    # Create CombinationMapper entry
                    mapper = CombinationMapper(
                        category_id=category_id,
                        equipment_id=equipment_id,
                        code_id=None,  # Will be set manually for specific codes
                        label_id=None  # Will be set manually for specific labels
                    )
                    self.db.add(mapper)
                    self.stats["mappings_created"] += 1
                    mapper_count += 1
                    
                    print(f"  ✓ Linked {equipment_name} → {category_name}")
                    
                    # Commit every 100 mappings
                    if mapper_count % 100 == 0:
                        self.db.commit()
                        batch_count += 1
                        print(f"\n  ✓ Committed batch {batch_count} ({mapper_count} mappings)")
                
                except Exception as e:
                    self.db.rollback()
                    self.stats["errors"].append(f"Row {row_idx}: {str(e)}")
                    print(f"  ✗ Error processing row {row_idx}: {str(e)}")
            
            # Final commit
            if mapper_count % 100 != 0:
                self.db.commit()
                if mapper_count > 0:
                    print(f"\n  ✓ Final commit ({mapper_count} mappings)")
            
            print("\n✓ Excel file processed successfully")
        
        except Exception as e:
            self.db.rollback()
            print(f"\n✗ ERROR: {str(e)}")
            raise

    def print_summary(self):
        """Print import summary statistics."""
        print("\n" + "="*70)
        print("IMPORT SUMMARY")
        print("="*70)
        
        print(f"\nCategories:")
        print(f"  Created: {self.stats['categories_created']}")
        print(f"  Skipped: {self.stats['categories_skipped']}")
        
        print(f"\nEquipment:")
        print(f"  Created: {self.stats['equipment_created']}")
        print(f"  Skipped: {self.stats['equipment_skipped']}")
        
        print(f"\nCombination Mappers:")
        print(f"  Created: {self.stats['mappings_created']}")
        print(f"  Skipped: {self.stats['mappings_skipped']}")
        
        print(f"\nTotal Errors: {self.stats['errors'].__len__()}")
        
        if self.stats["errors"]:
            print("\nFirst 5 Errors:")
            for error in self.stats["errors"][:5]:
                print(f"  - {error}")
            if len(self.stats["errors"]) > 5:
                print(f"  ... and {len(self.stats['errors']) - 5} more errors")
        
        print("\n" + "="*70)
        print("✓ EQUIPMENT, CATEGORY & COMBINATION MAPPER IMPORT COMPLETE!")
        print("="*70)

    def run(self):
        """Run the complete import process."""
        try:
            self.import_equipment_and_categories()
            self.print_summary()
        except Exception as e:
            print(f"\n✗ IMPORT FAILED: {str(e)}")
            self.db.rollback()
            raise
        finally:
            self.db.close()


def main():
    """Main entry point for the script."""
    importer = EquipmentCategoryImporter()
    importer.run()


if __name__ == "__main__":
    main()
