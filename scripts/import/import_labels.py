#!/usr/bin/env python
"""
Master Data Import Script - Load Labels, Categories, and Equipment from Excel
This script imports data from two Excel files:
1. Illumine-i X LabelFriday master data.xlsx - Label data
2. Label filter data.xlsx - Category/Equipment/Combination mapping

Usage:
    python scripts/import/import_labels.py

Note: Ensure you are in the project root directory when running this script.
"""

import os
import sys
from pathlib import Path

# Add project root to path for absolute imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session
from app.core.database import engine, SessionLocal
from app.models.label import Label
from app.models.category import Category
from app.models.equipment import Equipment
from app.models.combination_mapper import CombinationMapper
from app.models.code import Code


class MasterDataImporter:
    """
    Handles importing master data from Excel files to PostgreSQL.
    - Validates data before inserting
    - Handles duplicates gracefully
    - Provides detailed logging
    """

    def __init__(self):
        self.db = SessionLocal()
        self.stats = {
            "labels_created": 0,
            "labels_skipped": 0,
            "categories_created": 0,
            "categories_skipped": 0,
            "equipment_created": 0,
            "equipment_skipped": 0,
            "mappings_created": 0,
            "mappings_skipped": 0,
            "errors": []
        }

    def get_excel_path(self, filename: str) -> str:
        """
        Build absolute path to Excel file in data/master/ or app/data/master/
        Supports both running from project root and from scripts/import directory.
        """
        possible_paths = [
            PROJECT_ROOT / "data" / "master" / filename,
            PROJECT_ROOT / "app" / "data" / "master" / filename,
            Path.cwd() / "data" / "master" / filename,
            Path.cwd() / "app" / "data" / "master" / filename,
        ]

        for path in possible_paths:
            if path.exists():
                return str(path)

        raise FileNotFoundError(
            f"Excel file '{filename}' not found at:\n" +
            "\n".join(f"  - {p}" for p in possible_paths)
        )

    def load_workbook_safely(self, filepath: str):
        """Load Excel workbook with error handling."""
        try:
            return openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            raise Exception(f"Failed to load {filepath}: {str(e)}")

    def import_labels(self):
        """
        Import labels from 'Illumine-i X LabelFriday master data.xlsx'
        Expected columns: upc_code, label_number, label_name, name, field_type, description
        """
        print("\n" + "="*60)
        print("IMPORTING LABELS")
        print("="*60)

        try:
            filepath = self.get_excel_path("Illumine-i X LabelFriday master data.xlsx")
            print(f"Loading: {filepath}")

            wb = self.load_workbook_safely(filepath)
            ws = wb.active

            # Get headers (first row)
            headers = [cell.value for cell in ws[1]]
            print(f"Columns found: {headers}\n")

            # Map headers to column indices
            try:
                upc_idx = headers.index("upc_code") if "upc_code" in headers else 0
                label_number_idx = headers.index("label_number") if "label_number" in headers else 1
                label_name_idx = headers.index("label_name") if "label_name" in headers else 2
                name_idx = headers.index("name") if "name" in headers else 3
                field_type_idx = headers.index("field_type") if "field_type" in headers else 4
                description_idx = headers.index("description") if "description" in headers else 5
            except (ValueError, IndexError) as e:
                self.stats["errors"].append(f"Header mapping error: {str(e)}")
                print(f"WARNING: Could not map all headers, using default positions")

            # Iterate through data rows (skip header row 1)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:  # Skip empty rows
                        continue

                    upc_code = str(row[upc_idx]).strip() if row[upc_idx] else None
                    label_number = str(row[label_number_idx]).strip() if row[label_number_idx] else None
                    label_name = str(row[label_name_idx]).strip() if row[label_name_idx] else None
                    name = str(row[name_idx]).strip() if row[name_idx] else None
                    field_type = str(row[field_type_idx]).strip() if row[field_type_idx] else None
                    description = str(row[description_idx]).strip() if row[description_idx] else None

                    # Validate required fields
                    if not label_number:
                        self.stats["errors"].append(f"Row {row_idx}: Missing label_number")
                        self.stats["labels_skipped"] += 1
                        continue

                    # Check for duplicates
                    existing = self.db.query(Label).filter(
                        Label.label_number == label_number
                    ).first()

                    if existing:
                        self.stats["labels_skipped"] += 1
                        print(f"  SKIP: Label '{label_number}' already exists")
                        continue

                    # Create label
                    label = Label(
                        upc_code=upc_code,
                        label_number=label_number,
                        label_name=label_name or f"Label {label_number}",
                        name=name,
                        field_type=field_type,
                        description=description,
                        is_active=True
                    )
                    self.db.add(label)
                    self.db.commit()
                    self.stats["labels_created"] += 1
                    print(f"  CREATE: Label '{label_number}' (UPC: {upc_code})")

                except Exception as e:
                    self.db.rollback()
                    self.stats["errors"].append(f"Row {row_idx}: {str(e)}")
                    self.stats["labels_skipped"] += 1
                    print(f"  ERROR: Row {row_idx} - {str(e)}")

            print(f"\nLabels Summary:")
            print(f"  Created: {self.stats['labels_created']}")
            print(f"  Skipped: {self.stats['labels_skipped']}")

        except Exception as e:
            self.stats["errors"].append(f"Labels import failed: {str(e)}")
            print(f"\nERROR: {str(e)}")

    def import_category_equipment_mapping(self):
        """
        Import categories, equipment, and combinations from 'Label filter data.xlsx'
        Expected columns: category, equipment, label_number, code_name (optional)
        """
        print("\n" + "="*60)
        print("IMPORTING CATEGORIES, EQUIPMENT & COMBINATIONS")
        print("="*60)

        try:
            filepath = self.get_excel_path("Label filter data.xlsx")
            print(f"Loading: {filepath}")

            wb = self.load_workbook_safely(filepath)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]
            print(f"Columns found: {headers}\n")

            # Map headers to column indices
            try:
                category_idx = headers.index("category") if "category" in headers else 0
                equipment_idx = headers.index("equipment") if "equipment" in headers else 1
                label_number_idx = headers.index("label_number") if "label_number" in headers else 2
                code_name_idx = headers.index("code_name") if "code_name" in headers else 3
            except (ValueError, IndexError):
                print(f"WARNING: Could not map all headers, using default positions")

            # Iterate through data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row or not row[0]:  # Skip empty rows
                        continue

                    category_name = str(row[category_idx]).strip() if row[category_idx] else None
                    equipment_name = str(row[equipment_idx]).strip() if row[equipment_idx] else None
                    label_number = str(row[label_number_idx]).strip() if row[label_number_idx] else None
                    code_name = str(row[code_name_idx]).strip() if row[code_name_idx] else None

                    # At least label is required
                    if not label_number:
                        self.stats["errors"].append(f"Row {row_idx}: Missing label_number")
                        continue

                    # Get or create category
                    category_id = None
                    if category_name:
                        category = self.db.query(Category).filter(
                            Category.name == category_name
                        ).first()
                        if not category:
                            category = Category(name=category_name)
                            self.db.add(category)
                            self.db.commit()
                            self.stats["categories_created"] += 1
                            print(f"  CREATE: Category '{category_name}'")
                        else:
                            self.stats["categories_skipped"] += 1
                        category_id = category.id

                    # Get or create equipment
                    equipment_id = None
                    if equipment_name:
                        equipment = self.db.query(Equipment).filter(
                            Equipment.name == equipment_name
                        ).first()
                        if not equipment:
                            equipment = Equipment(name=equipment_name)
                            self.db.add(equipment)
                            self.db.commit()
                            self.stats["equipment_created"] += 1
                            print(f"  CREATE: Equipment '{equipment_name}'")
                        else:
                            self.stats["equipment_skipped"] += 1
                        equipment_id = equipment.id

                    # Get label
                    label = self.db.query(Label).filter(
                        Label.label_number == label_number
                    ).first()
                    if not label:
                        self.stats["errors"].append(
                            f"Row {row_idx}: Label '{label_number}' not found"
                        )
                        continue

                    # Get code (optional)
                    code_id = None
                    if code_name:
                        code = self.db.query(Code).filter(
                            Code.code_name == code_name
                        ).first()
                        if code:
                            code_id = code.id
                        else:
                            self.stats["errors"].append(
                                f"Row {row_idx}: Code '{code_name}' not found (using label without code)"
                            )

                    # Create combination mapping
                    # Only create if not all nulls
                    if category_id or equipment_id or code_id:
                        existing_combo = self.db.query(CombinationMapper).filter(
                            CombinationMapper.label_id == label.id,
                            CombinationMapper.category_id == category_id,
                            CombinationMapper.equipment_id == equipment_id,
                            CombinationMapper.code_id == code_id,
                        ).first()

                        if existing_combo:
                            self.stats["mappings_skipped"] += 1
                            print(f"  SKIP: Combination already exists (Label: {label_number})")
                        else:
                            mapper = CombinationMapper(
                                category_id=category_id,
                                equipment_id=equipment_id,
                                code_id=code_id,
                                label_id=label.id
                            )
                            self.db.add(mapper)
                            self.db.commit()
                            self.stats["mappings_created"] += 1
                            combo_desc = f"Cat: {category_name or 'N/A'}, Eq: {equipment_name or 'N/A'}"
                            print(f"  CREATE: Combination for Label '{label_number}' ({combo_desc})")

                except Exception as e:
                    self.db.rollback()
                    self.stats["errors"].append(f"Row {row_idx}: {str(e)}")
                    self.stats["mappings_skipped"] += 1
                    print(f"  ERROR: Row {row_idx} - {str(e)}")

            print(f"\nCategories/Equipment Summary:")
            print(f"  Categories Created: {self.stats['categories_created']}")
            print(f"  Equipment Created: {self.stats['equipment_created']}")
            print(f"  Combinations Created: {self.stats['mappings_created']}")
            print(f"  Mappings Skipped: {self.stats['mappings_skipped']}")

        except Exception as e:
            self.stats["errors"].append(f"Category/Equipment import failed: {str(e)}")
            print(f"\nERROR: {str(e)}")

    def print_final_report(self):
        """Print comprehensive import report."""
        print("\n" + "="*60)
        print("IMPORT COMPLETE - FINAL REPORT")
        print("="*60)

        print(f"\nLabels:")
        print(f"  Created: {self.stats['labels_created']}")
        print(f"  Skipped: {self.stats['labels_skipped']}")

        print(f"\nCategories:")
        print(f"  Created: {self.stats['categories_created']}")
        print(f"  Skipped: {self.stats['categories_skipped']}")

        print(f"\nEquipment:")
        print(f"  Created: {self.stats['equipment_created']}")
        print(f"  Skipped: {self.stats['equipment_skipped']}")

        print(f"\nCombinations:")
        print(f"  Created: {self.stats['mappings_created']}")
        print(f"  Skipped: {self.stats['mappings_skipped']}")

        if self.stats["errors"]:
            print(f"\nErrors ({len(self.stats['errors'])}):")
            for err in self.stats["errors"][:10]:  # Show first 10
                print(f"  - {err}")
            if len(self.stats["errors"]) > 10:
                print(f"  ... and {len(self.stats['errors']) - 10} more")

        print("\n" + "="*60)

    def run(self):
        """Execute full import workflow."""
        try:
            print("\n" + "="*60)
            print("AHJ ENGINE - MASTER DATA IMPORT")
            print("="*60)
            print(f"Project Root: {PROJECT_ROOT}")
            print(f"Database: {os.getenv('DATABASE_URL', 'postgresql://localhost/ahj_db')}")

            self.import_labels()
            self.import_category_equipment_mapping()
            self.print_final_report()

        except Exception as e:
            print(f"\nFATAL ERROR: {str(e)}")
            sys.exit(1)
        finally:
            self.db.close()


if __name__ == "__main__":
    importer = MasterDataImporter()
    importer.run()
