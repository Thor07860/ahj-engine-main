#!/usr/bin/env python
"""
Code Import Script - Load Codes from Excel
This script imports code data from the "Code Mapping" sheet in data/code_mapping.xlsx

The Excel file should contain category blocks with columns:
- Local Code
- Related Standard Code

For each code, the script:
1. Determines the code type, amendment type, and issuing body
2. Extracts the edition year from the title
3. Creates CodeType and CodeAmendment entries if missing
4. Checks for duplicates before inserting
5. Commits changes in batches

Usage:
    python scripts/import/import_codes.py

Note: Ensure you are in the project root directory when running this script.
"""

import os
import sys
import re
from pathlib import Path
from typing import Optional, Tuple, Dict, List

# Add project root to path for absolute imports
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

from sqlalchemy.orm import Session
from sqlalchemy import and_

from app.core.database import engine, SessionLocal
from app.models.code import Code
from app.models.code_type import CodeType
from app.models.code_amendment import CodeAmendment
from app.models.applicable_code_category import ApplicableCodeCategory


class CodeImporter:
    """
    Handles importing code data from Excel to PostgreSQL.
    - Validates and parses code entries
    - Creates reference data automatically
    - Handles duplicates gracefully
    - Provides detailed logging and statistics
    """

    def __init__(self):
        self.db: Session = SessionLocal()
        self.stats = {
            "codes_created": 0,
            "codes_skipped": 0,
            "codes_errors": 0,
            "by_category": {},
            "errors": []
        }
        
        # Initialize category mapping
        self.category_map = {
            "Fire Code": "Fire",
            "Building Code": "Building",
            "Electrical Code": "Electrical",
            "Residential Code": "Residential",
            "Mechanical Code": "Mechanical",
            "Plumbing Code": "Plumbing",
            "Energy Code": "Energy",
        }
        
        # Reference data caches
        self.code_types_cache: Dict[str, int] = {}
        self.amendments_cache: Dict[str, int] = {}
        self.categories_cache: Dict[str, int] = {}

    def get_excel_path(self, filename: str) -> str:
        """
        Build absolute path to Excel file in data/ or app/data/ directory.
        Supports both running from project root and from scripts/import directory.
        """
        possible_paths = [
            PROJECT_ROOT / "data" / "master" / filename,
            PROJECT_ROOT / "app" / "data" / "master" / filename,
            PROJECT_ROOT / "data" / filename,
            PROJECT_ROOT / "app" / "data" / filename,
            Path.cwd() / "data" / "master" / filename,
            Path.cwd() / "app" / "data" / "master" / filename,
            Path.cwd() / "data" / filename,
            Path.cwd() / "app" / "data" / filename,
        ]

        for path in possible_paths:
            if path.exists():
                print(f"✓ Found Excel file: {path}")
                return str(path)

        raise FileNotFoundError(
            f"Excel file '{filename}' not found at:\n" +
            "\n".join(f"  - {p}" for p in possible_paths)
        )

    def load_workbook_safely(self, filepath: str):
        """Load Excel workbook with error handling."""
        try:
            return openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            raise Exception(f"Failed to load {filepath}: {str(e)}")

    def ensure_reference_data(self):
        """
        Create all reference data (CodeTypes, CodeAmendments, Categories) if missing.
        """
        print("\n--- Ensuring Reference Data ---")
        
        # CodeAmendments
        amendments = ["LOCAL", "STANDARD", "STATE", "INTERNATIONAL"]
        for amendment_name in amendments:
            existing = self.db.query(CodeAmendment).filter_by(name=amendment_name).first()
            if not existing:
                amendment = CodeAmendment(name=amendment_name)
                self.db.add(amendment)
                print(f"  Created CodeAmendment: {amendment_name}")
            else:
                existing_id = existing.id
                self.amendments_cache[amendment_name] = existing_id
        
        self.db.commit()
        
        # Reload cache
        for amendment_name in amendments:
            amendment = self.db.query(CodeAmendment).filter_by(name=amendment_name).first()
            if amendment:
                self.amendments_cache[amendment_name] = amendment.id
        
        # ApplicableCodeCategories
        categories = ["Electrical", "Building", "Fire", "Residential", "Mechanical", "Plumbing", "Energy"]
        for category_name in categories:
            existing = self.db.query(ApplicableCodeCategory).filter_by(name=category_name).first()
            if not existing:
                category = ApplicableCodeCategory(name=category_name)
                self.db.add(category)
                print(f"  Created ApplicableCodeCategory: {category_name}")
            else:
                self.categories_cache[category_name] = existing.id
        
        self.db.commit()
        
        # Reload cache
        for category_name in categories:
            category = self.db.query(ApplicableCodeCategory).filter_by(name=category_name).first()
            if category:
                self.categories_cache[category_name] = category.id
        
        print(f"✓ Reference data ready")

    def extract_year(self, title: str) -> Optional[str]:
        """
        Extract the first 4-digit year from the title.
        Example: "2021 International Building Code" → "2021"
        """
        if not title:
            return None
        
        match = re.search(r'\b(20\d{2}|19\d{2})\b', title)
        return match.group(1) if match else None

    def detect_issuing_body(self, title: str) -> Optional[str]:
        """
        Detect the issuing body from the code title.
        Priority:
        1. "International" in title → ICC
        2. "National Electrical Code" or "NFPA" → NFPA
        3. Otherwise → NULL
        """
        if not title:
            return None
        
        title_upper = title.upper()
        
        if "INTERNATIONAL" in title_upper:
            return "ICC"
        
        if "NATIONAL ELECTRICAL CODE" in title_upper or "NFPA" in title_upper:
            return "NFPA"
        
        return None

    def infer_code_type(self, title: str) -> str:
        """
        Infer the code type from the title.
        Returns the code type title (e.g., "Building Code", "Electrical Code").
        """
        if not title:
            return "General Code"
        
        title_upper = title.upper()
        
        # Check for specific code types
        if "ELECTRICAL" in title_upper or "NEC" in title_upper:
            return "Electrical Code"
        elif "BUILDING" in title_upper or "IBC" in title_upper:
            return "Building Code"
        elif "FIRE" in title_upper or "IFPC" in title_upper:
            return "Fire Code"
        elif "RESIDENTIAL" in title_upper or "IRC" in title_upper:
            return "Residential Code"
        elif "MECHANICAL" in title_upper or "IMC" in title_upper:
            return "Mechanical Code"
        elif "PLUMBING" in title_upper or "IPC" in title_upper:
            return "Plumbing Code"
        elif "ENERGY" in title_upper or "IECC" in title_upper:
            return "Energy Code"
        else:
            return "General Code"

    def get_or_create_code_type(self, code_type_title: str) -> int:
        """
        Get or create a CodeType by title.
        Returns the ID.
        """
        # Check cache first
        if code_type_title in self.code_types_cache:
            return self.code_types_cache[code_type_title]
        
        # Query database
        existing = self.db.query(CodeType).filter_by(title=code_type_title).first()
        
        if not existing:
            code_type = CodeType(title=code_type_title, key=code_type_title.replace(" ", "_").lower())
            self.db.add(code_type)
            self.db.flush()
            code_type_id = code_type.id
        else:
            code_type_id = existing.id
        
        # Cache it
        self.code_types_cache[code_type_title] = code_type_id
        return code_type_id

    def code_exists(self, title: str, amendment_id: int, category_id: Optional[int]) -> bool:
        """
        Check if a code already exists with the same title, amendment type, and category.
        """
        query = self.db.query(Code).filter(
            and_(
                Code.title == title,
                Code.code_amendments == amendment_id,
                Code.applicable_code_category_id == category_id
            )
        )
        return query.first() is not None

    def process_code_entry(self, code_title: str, amendment_type: str, category_name: str) -> bool:
        """
        Process a single code entry from Excel.
        
        Args:
            code_title: The code name/title from Excel
            amendment_type: "LOCAL" or "STANDARD"
            category_name: The category (e.g., "Fire", "Electrical")
        
        Returns:
            True if code was created, False if skipped or error
        """
        if not code_title or not code_title.strip():
            return False
        
        code_title = code_title.strip()
        
        try:
            # Initialize category stats if needed
            if category_name not in self.stats["by_category"]:
                self.stats["by_category"][category_name] = {"created": 0, "skipped": 0}
            
            # Get or create CodeAmendment
            amendment_id = self.amendments_cache.get(amendment_type)
            if not amendment_id:
                raise ValueError(f"Amendment type '{amendment_type}' not found")
            
            # Get or create ApplicableCodeCategory
            category_id = self.categories_cache.get(category_name)
            if not category_id:
                raise ValueError(f"Category '{category_name}' not found")
            
            # Check for duplicates
            if self.code_exists(code_title, amendment_id, category_id):
                self.stats["codes_skipped"] += 1
                self.stats["by_category"][category_name]["skipped"] += 1
                return False
            
            # Extract metadata
            edition = self.extract_year(code_title)
            issuing_body = self.detect_issuing_body(code_title)
            code_type_title = self.infer_code_type(code_title)
            code_type_id = self.get_or_create_code_type(code_type_title)
            
            # Create code entry
            code = Code(
                code_name=code_title[:255],  # code_name is required, use title
                title=code_title,
                code_type_id=code_type_id,
                code_amendments=amendment_id,
                applicable_code_category_id=category_id,
                edition=edition,
                issuing_body=issuing_body,
                description=None,
                state_id=None,
                adopted_at=None
            )
            
            self.db.add(code)
            self.stats["codes_created"] += 1
            self.stats["by_category"][category_name]["created"] += 1
            
            return True
        
        except Exception as e:
            self.stats["codes_errors"] += 1
            error_msg = f"Error processing '{code_title}': {str(e)}"
            self.stats["errors"].append(error_msg)
            print(f"  ✗ {error_msg}")
            return False

    def import_codes(self):
        """
        Import codes from the "Code Mapping" sheet in Excel.
        
        Sheet structure (HORIZONTAL):
        Row 1: Category headers (Electrical Code, Building Code, Fire Code, etc.)
        Row 2: Column headers (Local Code, Related standard code)
        Rows 3+: Actual code data
        
        Columns organized as: [Cat1 Local, Cat1 Standard, Cat2 Local, Cat2 Standard, ...]
        """
        print("\n" + "="*70)
        print("IMPORTING CODES FROM CODE MAPPING")
        print("="*70)
        
        # Ensure reference data exists
        self.ensure_reference_data()
        
        try:
            filepath = self.get_excel_path("code mapping.xlsx")
            print(f"\nLoading workbook: {filepath}")
            
            wb = self.load_workbook_safely(filepath)
            
            # Get the "Code Mapping" sheet
            if "Code Mapping" not in wb.sheetnames:
                raise ValueError(f"Sheet 'Code Mapping' not found. Available sheets: {wb.sheetnames}")
            
            ws = wb["Code Mapping"]
            print(f"✓ Found sheet: 'Code Mapping'")
            print(f"  Dimensions: {ws.dimensions}")
            
            # Parse the header structure
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            column_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            
            # Map columns to categories
            # Column pairs: (Local Code, Related Standard Code) under each category header
            column_mapping = {}  # column_index -> (category_name, is_standard)
            
            current_category = None
            for col_idx, category in enumerate(header_row[0] if header_row else []):
                if category and str(category).strip():
                    # Extract category name from header like "Electrical Code"
                    category_str = str(category).strip()
                    # Map to simplified name
                    for key, value in self.category_map.items():
                        if category_str.lower().startswith(key.lower()):
                            current_category = value
                            break
                    
                    if current_category:
                        # Next two columns should be Local Code and Related Standard Code
                        column_mapping[col_idx] = (current_category, False)  # Local Code
                        if col_idx + 1 < len(header_row[0]):
                            column_mapping[col_idx + 1] = (current_category, True)  # Standard Code
            
            if not column_mapping:
                print("⚠ No category headers found in row 1")
                return
            
            print(f"\n✓ Found {len(set(cat for cat, _ in column_mapping.values()))} categories")
            print("\n--- Processing Codes ---")
            
            # Process data rows (starting from row 3)
            code_count = 0
            batch_count = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
                row_values = [cell.value for cell in row]
                
                # Process each column pair (Local Code, Standard Code)
                for col_idx, (category_name, is_standard) in column_mapping.items():
                    if col_idx >= len(row_values):
                        continue
                    
                    code_title = row_values[col_idx]
                    amendment_type = "STANDARD" if is_standard else "LOCAL"
                    
                    # Try to process this code entry
                    if self.process_code_entry(code_title, amendment_type, category_name):
                        code_count += 1
                    
                    # Commit every 200 codes
                    if code_count % 200 == 0 and code_count > 0:
                        self.db.commit()
                        batch_count += 1
                        print(f"  ✓ Committed batch {batch_count} ({code_count} codes)")
            
            # Final commit
            if code_count % 200 != 0:
                self.db.commit()
                if code_count > 0:
                    print(f"  ✓ Final commit ({code_count} codes)")
            
            print("\n✓ Excel file processed successfully")
        
        except Exception as e:
            self.db.rollback()
            print(f"\n✗ ERROR: {str(e)}")
            raise

    def print_summary(self):
        """Print import summary statistics."""
        print("\n" + "="*70)
        print("CODE IMPORT SUMMARY")
        print("="*70)
        
        print(f"\nTotal Codes Created: {self.stats['codes_created']}")
        print(f"Total Codes Skipped (Duplicates): {self.stats['codes_skipped']}")
        print(f"Total Errors: {self.stats['codes_errors']}")
        
        print("\nBy Category:")
        for category, stats in sorted(self.stats["by_category"].items()):
            print(f"  {category:15} | Created: {stats['created']:4} | Skipped: {stats['skipped']:4}")
        
        if self.stats["errors"]:
            print("\nErrors:")
            for error in self.stats["errors"][:10]:  # Show first 10 errors
                print(f"  - {error}")
            if len(self.stats["errors"]) > 10:
                print(f"  ... and {len(self.stats['errors']) - 10} more errors")
        
        print("\n" + "="*70)
        print("CODE IMPORT COMPLETE!")
        print("="*70)

    def run(self):
        """Run the complete import process."""
        try:
            self.import_codes()
            self.print_summary()
        except Exception as e:
            print(f"\n✗ IMPORT FAILED: {str(e)}")
            self.db.rollback()
            raise
        finally:
            self.db.close()


def main():
    """Main entry point for the script."""
    importer = CodeImporter()
    importer.run()


if __name__ == "__main__":
    main()
